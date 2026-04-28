"""Parsers de fichiers Excel d'import d'écritures.

Format attendu (modèle Achats / Ventes / OD) :
    Date | Référence | Libellé | Compte | Débit | Crédit | Tiers (optionnel)

Toutes les lignes ayant la même `Référence` constituent une seule pièce qui
doit être équilibrée (R1).
"""
from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from apps.comptabilite.models import CompteComptable, Exercice, Journal
from apps.tiers.models import Tiers

COLONNES_ATTENDUES = ["Date", "Référence", "Libellé", "Compte", "Débit", "Crédit", "Tiers"]


def _parse_decimal(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0.00")
    try:
        return Decimal(str(v).replace(",", ".").replace(" ", ""))
    except (InvalidOperation, TypeError):
        raise ValueError(f"Montant invalide : {v!r}")


def _parse_date(v) -> date:
    if isinstance(v, date):
        return v
    raise ValueError(f"Date invalide ou absente : {v!r}")


def parse_excel_pieces(fichier_path: str, exercice: Exercice, journal: Journal) -> tuple[list[dict], list[dict]]:
    """Parse le fichier et retourne (pieces_a_creer, erreurs).

    Aucune écriture en base ici — uniquement validation et construction des dicts.
    """
    wb = load_workbook(fichier_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [{"ligne": 0, "motif": "Fichier vide"}]

    headers = [str(c).strip() if c else "" for c in rows[0]]
    if headers[: len(COLONNES_ATTENDUES) - 1] != COLONNES_ATTENDUES[:-1]:
        return [], [{
            "ligne": 1,
            "motif": f"Colonnes attendues : {COLONNES_ATTENDUES}, trouvées : {headers}",
        }]

    erreurs: list[dict] = []
    pieces_par_ref: dict[str, list[dict]] = defaultdict(list)

    for idx, row in enumerate(rows[1:], start=2):
        if all(c is None or c == "" for c in row):
            continue
        try:
            d = _parse_date(row[0])
            ref = str(row[1]).strip() if row[1] else ""
            libelle = str(row[2] or "").strip()
            compte_num = str(row[3] or "").strip()
            debit = _parse_decimal(row[4])
            credit = _parse_decimal(row[5])
            tiers_code = str(row[6]).strip() if len(row) > 6 and row[6] else ""

            if not ref:
                raise ValueError("Référence vide (sert à regrouper les lignes en pièce)")
            if not (exercice.date_debut <= d <= exercice.date_fin):
                raise ValueError(f"Date {d} hors exercice {exercice.code} (R5)")
            try:
                compte = CompteComptable.objects.get(numero=compte_num)
            except CompteComptable.DoesNotExist:
                raise ValueError(f"Compte {compte_num} inexistant (R6)")
            if (debit > 0) == (credit > 0):
                raise ValueError("Une ligne doit être soit débit soit crédit, pas les deux ni zéro")
            tiers = None
            if compte.collectif_tiers:
                if not tiers_code:
                    raise ValueError(f"Compte {compte_num} collectif : code tiers obligatoire (R7)")
                try:
                    tiers = Tiers.objects.get(code_auxiliaire=tiers_code)
                except Tiers.DoesNotExist:
                    raise ValueError(f"Tiers {tiers_code} inexistant")
            elif tiers_code:
                raise ValueError(f"Compte {compte_num} non collectif : tiers interdit (R7)")

            pieces_par_ref[ref].append({
                "date_piece": d, "reference": ref, "libelle": libelle,
                "compte": compte, "tiers": tiers, "debit": debit, "credit": credit,
            })
        except Exception as e:
            erreurs.append({"ligne": idx, "motif": str(e)})

    pieces = []
    for ref, lignes in pieces_par_ref.items():
        total_d = sum((row["debit"] for row in lignes), Decimal("0.00"))
        total_c = sum((row["credit"] for row in lignes), Decimal("0.00"))
        if total_d != total_c:
            erreurs.append({
                "ligne": "—", "motif": f"Pièce {ref} déséquilibrée (D={total_d} / C={total_c}) (R1)",
            })
            continue
        pieces.append({
            "journal": journal, "exercice": exercice,
            "date_piece": lignes[0]["date_piece"], "reference": ref,
            "libelle": lignes[0]["libelle"] or f"Import {ref}",
            "lignes": lignes,
        })

    return pieces, erreurs
