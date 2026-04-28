"""Génération de fichiers Excel via openpyxl en mode write_only (streaming)."""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.etats import selectors

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2D6A4F")
NUM_FORMAT = "#,##0.00"


def _new_workbook(sheet_title: str):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=sheet_title[:31])
    return wb, ws


def _to_bytes(wb) -> bytes:
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _header_cell(value: str):
    from openpyxl.cell import WriteOnlyCell

    cell = WriteOnlyCell(None, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    return cell


def _num_cell(value):
    from openpyxl.cell import WriteOnlyCell

    cell = WriteOnlyCell(None, value=value)
    cell.number_format = NUM_FORMAT
    cell.alignment = Alignment(horizontal="right")
    return cell


def build_balance_xlsx(exercice, date_debut=None, date_fin=None, classes=None) -> bytes:
    wb, ws = _new_workbook(f"Balance {exercice.code}")
    headers = ["Compte", "Libellé", "Classe", "Sens", "Débit", "Crédit", "Solde"]
    ws.append([_header_cell(h) for h in headers])
    total_d = total_c = 0
    for row in selectors.balance(exercice, date_debut, date_fin, classes):
        ws.append([
            row["compte__numero"],
            row["compte__libelle"],
            row["compte__classe"],
            row["compte__sens"],
            _num_cell(row["total_debit"] or 0),
            _num_cell(row["total_credit"] or 0),
            _num_cell(row["solde"] or 0),
        ])
        total_d += row["total_debit"] or 0
        total_c += row["total_credit"] or 0
    ws.append(["TOTAL", "", "", "", _num_cell(total_d), _num_cell(total_c), _num_cell(total_d - total_c)])
    return _to_bytes(wb)


def build_grand_livre_xlsx(compte, exercice, date_debut=None, date_fin=None) -> bytes:
    wb, ws = _new_workbook(f"GL {compte.numero}")
    headers = ["Date", "Journal", "N° Pièce", "Tiers", "Libellé", "Débit", "Crédit", "Solde", "Lettre"]
    ws.append([_header_cell(h) for h in headers])
    for r in selectors.grand_livre_compte(compte, exercice, date_debut, date_fin):
        ws.append([
            r["date"].isoformat() if r["date"] else "",
            r["journal"],
            r["piece_numero"] or "",
            r["tiers"].code_auxiliaire if r["tiers"] else "",
            r["libelle"],
            _num_cell(r["debit"] or 0),
            _num_cell(r["credit"] or 0),
            _num_cell(r["solde"]),
            r["lettre"],
        ])
    return _to_bytes(wb)


def build_journal_xlsx(journal_obj, exercice, date_debut=None, date_fin=None) -> bytes:
    wb, ws = _new_workbook(f"Journal {journal_obj.code}")
    headers = ["Date", "N°", "Référence", "Compte", "Tiers", "Libellé", "Débit", "Crédit"]
    ws.append([_header_cell(h) for h in headers])
    for piece in selectors.journal(journal_obj, exercice, date_debut, date_fin):
        for ligne in piece.lignes.all():
            ws.append([
                piece.date_piece.isoformat(),
                piece.numero or "",
                piece.reference,
                ligne.compte.numero,
                ligne.tiers.code_auxiliaire if ligne.tiers else "",
                ligne.libelle or piece.libelle,
                _num_cell(ligne.debit or 0),
                _num_cell(ligne.credit or 0),
            ])
    return _to_bytes(wb)
