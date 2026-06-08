import csv
from datetime import date, datetime
from decimal import Decimal
from io import TextIOWrapper

from django.db import transaction
from openpyxl import load_workbook

from apps.comptabilite.models import LigneEcriture

from .matching import choisir_ecriture
from .models import LigneReleve, ReleveBancaire


def _parse_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()


@transaction.atomic
def creer_releve_depuis_lignes(compte_bancaire, lignes: list[dict], *, date_debut, date_fin,
                               solde_initial=Decimal("0.00"), solde_final=Decimal("0.00"),
                               fichier_source=None) -> ReleveBancaire:
    """Crée un ReleveBancaire + ses LigneReleve à partir d'une liste de dicts
    {date_operation, libelle, montant (signé), reference_banque}.
    """
    releve = ReleveBancaire.objects.create(
        compte_bancaire=compte_bancaire, date_debut=date_debut, date_fin=date_fin,
        solde_initial=solde_initial, solde_final=solde_final, fichier_source=fichier_source,
    )
    objets = [
        LigneReleve(
            releve=releve, date_operation=ligne["date_operation"], libelle=ligne.get("libelle", ""),
            montant=ligne["montant"], reference_banque=ligne.get("reference_banque", ""),
        )
        for ligne in lignes
    ]
    LigneReleve.objects.bulk_create(objets)
    return releve


def lire_lignes_excel(fichier_path) -> list[dict]:
    """Lit un relevé Excel : en-têtes (1re ligne) date, libelle, montant, reference."""
    wb = load_workbook(fichier_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    entetes = [str(h).strip().lower() if h else "" for h in next(rows)]
    idx = {nom: i for i, nom in enumerate(entetes)}
    lignes = []
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        lignes.append({
            "date_operation": _parse_date(row[idx["date"]]),
            "libelle": str(row[idx["libelle"]] or ""),
            "montant": Decimal(str(row[idx["montant"]])),
            "reference_banque": str(row[idx["reference"]] or "") if "reference" in idx else "",
        })
    return lignes


def lire_lignes_csv(fichier) -> list[dict]:
    """Lit un relevé CSV (séparateur ';' ou ','), colonnes date, libelle, montant, reference."""
    if hasattr(fichier, "read"):
        wrapper = TextIOWrapper(fichier, encoding="utf-8-sig")
    else:
        wrapper = open(fichier, encoding="utf-8-sig")  # noqa: SIM115
    sample = wrapper.read(2048)
    wrapper.seek(0)
    dialect = csv.Sniffer().sniff(sample, delimiters=";,")
    reader = csv.DictReader(wrapper, dialect=dialect)
    lignes = []
    for row in reader:
        row = {(k or "").strip().lower(): v for k, v in row.items()}
        lignes.append({
            "date_operation": _parse_date(row["date"]),
            "libelle": row.get("libelle", ""),
            "montant": Decimal(str(row["montant"]).replace(" ", "").replace(",", ".")),
            "reference_banque": row.get("reference", ""),
        })
    return lignes


@transaction.atomic
def pointer_manuellement(ligne_releve, ligne_ecriture):
    """Lie manuellement une ligne de relevé à une écriture. Refuse si déjà pointée
    ou si l'écriture n'est pas sur le compte bancaire du relevé.
    """
    compte = ligne_releve.releve.compte_bancaire.compte_comptable_id
    if ligne_ecriture.compte_id != compte:
        raise ValueError("L'écriture n'appartient pas au compte bancaire du relevé")
    # Re-fetch with lock to get the current DB state and prevent race conditions
    ecriture_db = LigneEcriture.objects.select_for_update().get(pk=ligne_ecriture.pk)
    if ecriture_db.pointee:
        raise ValueError("Écriture déjà pointée")
    LigneEcriture.objects.filter(pk=ligne_ecriture.pk).update(pointee=True)
    ligne_releve.ligne_ecriture_pointee = ligne_ecriture
    ligne_releve.statut = "POINTEE_MANUEL"
    ligne_releve.save(update_fields=["ligne_ecriture_pointee", "statut"])


@transaction.atomic
def depointer(ligne_releve):
    """Annule le pointage d'une ligne de relevé."""
    if ligne_releve.ligne_ecriture_pointee_id:
        LigneEcriture.objects.filter(pk=ligne_releve.ligne_ecriture_pointee_id).update(pointee=False)
    ligne_releve.ligne_ecriture_pointee = None
    ligne_releve.statut = "NON_POINTEE"
    ligne_releve.save(update_fields=["ligne_ecriture_pointee", "statut"])


@transaction.atomic
def pointer_automatiquement(releve, fenetre_jours: int = 5) -> int:
    """Pointe les LigneReleve NON_POINTEE avec les écritures du compte bancaire.
    Retourne le nombre de lignes pointées.
    """
    compte = releve.compte_bancaire.compte_comptable
    n = 0
    for ligne in releve.lignes.filter(statut="NON_POINTEE"):
        candidats = [
            {
                "id": e.id, "debit": e.debit, "credit": e.credit,
                "date": e.date_operation or e.piece.date_piece, "libelle": e.libelle,
            }
            for e in LigneEcriture.objects.select_related("piece").filter(
                compte=compte, pointee=False, piece__statut="VALIDEE"
            )
        ]
        choix = choisir_ecriture(ligne.montant, ligne.date_operation, ligne.libelle, candidats, fenetre_jours)
        if choix is None:
            continue
        LigneEcriture.objects.filter(pk=choix).update(pointee=True)
        ligne.ligne_ecriture_pointee_id = choix
        ligne.statut = "POINTEE_AUTO"
        ligne.save(update_fields=["ligne_ecriture_pointee", "statut"])
        n += 1
    return n
